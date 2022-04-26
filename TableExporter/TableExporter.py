#pip install pyinstaller
#pip install xlrd ## xlrd는 정책변경으로 xls 만 읽을 수 있어서 openpyxl 로 교체.
#pip install openpyxl

#import xlrd
import openpyxl

import sys
import os
import timeit
import enum


#------------------------------------------------------------------------
# 시간 체크.
#------------------------------------------------------------------------
timer_stack_ = list()
def BeginTimer(tag : str):
	start_time = timeit.default_timer()
	timer_stack_.append((tag, start_time))
def EndTimer() -> (str, float):
	stack_count = len(timer_stack_)
	if stack_count > 0:
		last_time = timer_stack_.pop()
		tag = str(last_time[0])
		start_time = float(last_time[1])
		end_time = timeit.default_timer()
		return (("\t" * (stack_count - 2)) + tag, end_time - start_time)
	return ("", 0.0)
def PrintTimer():
	print(end())
def PrintTimer(data : (str, float)):
	tag = data[0]
	pass_time = data[1]
	print(f"{tag} : {pass_time:.4f}s")


#------------------------------------------------------------------------
# C# 에서 사용중인 키워드 (전체).
# 팁 : 최소한의 검사용으로 작성 해놓았지만 C#의 버전업시 어떻게 바뀔지는 모르기 때문에...
# 가급적이면 필드명 등에서 대소문자를 적절히 혼용해서 키워드와의 중복을 피할 것.
#------------------------------------------------------------------------
csharp_strong_keyword_list_ = [ "abstract", "as", "base", "bool", "break", "byte", "case", "catch", "char", "checked", "class", "const", "delegate", "do", "double", "else", "enum", "event", "explicit", "extern", "false", "finally", "fixed", "float", "for", "foreach", "goto", "if", "implicit", "in", "int", "interface", "internal", "lock", "long", "namespace", "new", "null", "object", "function", "declare", "define", "out", "override", "params", "private", "protected", "public", "readonly", "ref", "return", "sbyte", "sealed", "short", "sizeof", "stackalloc", "static", "string", "struct", "switch", "this", "throw", "true", "try", "typeof", "uint", "ulong", "unchecked", "unsafe", "ushort", "using", "virtual", "void", "volatile", "while" ]


#------------------------------------------------------------------------
# C#에서 사용중인 키워드 (데이터타입제외).
#------------------------------------------------------------------------
csharp_weak_keyword_list_ = [ "abstract", "as", "base", "break", "byte", "case", "catch", "checked", "class", "const", "delegate", "do", "else", "enum", "event", "explicit", "extern", "false", "finally", "fixed", "for", "foreach", "goto", "if", "implicit", "in", "interface", "internal", "lock", "namespace", "new", "null", "object", "function", "declare", "define", "out", "override", "params", "private", "protected", "public", "readonly", "ref", "return", "sealed", "sizeof", "stackalloc", "static", "struct", "switch", "this", "throw", "true", "try", "typeof", "unchecked", "unsafe", "using", "virtual", "void", "volatile", "while" ]


class DataKind(enum.Enum):
	UNKNOWN = enum.auto()
	BOOLEAN = enum.auto()
	NUMBER = enum.auto()
	REAL = enum.auto()
	TEXT = enum.auto()


def GetAbsPath(path : str = "") -> str:
	if os.path.isfile(path) or os.path.isdir(path):
		return os.path.abspath(path)

	if getattr(sys, 'frozen', False):
		return os.path.abspath(os.path.dirname(sys.executable) + path) # 응용프로그램 경로.
	elif __file__: # 현재 스크립트파일의 경로.
		return os.path.abspath(os.path.dirname(__file__) + path)


def IsCorrectName(value : str):
	if not value:
		return False

	# 알파벳 ==> 숫자의 순으로 호출되어야하며 알파벳과 숫자 외의 문자열은 실패.
	is_correct = False
	for val in value:
		if val.isalpha():
			is_correct = True
		elif val.isnumeric():
			if not is_correct:
				return False
		elif val == "_":
			is_correct = True
		else:
			return False
	return True



def DataValueToJsonValue(value : str, kind : DataKind, is_array : bool):
	if kind == DataKind.TEXT:
		return f"\"{value}\""
	elif is_array:
		result = str()
		result += "["
		val_list = value.split(",")
		val_list_count = len(val_list)
		for val_index in range(val_list_count):
			result += f"{field_value}"
			if val_index + 1 < val_list_count:
				result += ", "
		result += "]"
	else:
		return f"{value}"

class NativeTable:
	name_ : str
	row_list_ : list # row[col[]]
	def __init__(self):
		self.name_ = str()
		self.row_list_ = list()


class Field:
	name_  : str
	type_ : str
	kind_ : DataKind
	is_array_ : bool
	comment_ : str
	data_ : list
	def __init__(self):
		self.name_ = str()
		self.type_ = str()
		self.kind_ = DataKind.UNKNOWN
		self.is_array_ = bool()
		self.comment_ = str()
		self.data_ = list()

class Table:
	name_ : str
	field_dict_ : dict # key:index, value:Field
	data_count_ : int
	def __init__(self):
		self.name_ = str()
		field_dict_ = dict()
		data_count_ = int()


def CreateNativeTableListFromXlsxFile(file_name : str, startswith = "#") -> list:
	nativetable_list = list()
	#xlrd은 더이상 사용하지 않음.
	#book = xlrd.open_workbook(file_name)
	#for sheet in book.sheets():
	#	if sheet.name.startswith(startswith):
	#		native_table = NativeTable()
	#		native_table.name_ = sheet.name.replace(startswith, "")
	#		for row_index in range(sheet.nrows):
	#			cell_list = sheet.row_values(row_index)
	#			cols = list()
	#			for cell in cell_list:
	#				cols.append(str(cell))
	#			native_table.row_list_.append(cols)
	#		native_table_list.append(native_table)
	#return native_table_list
	book = openpyxl.load_workbook(file_name)
	for sheetname in book.sheetnames:
		if not sheetname.startswith(startswith):
			continue;
		sheet = book[sheetname]
		if not sheet:
			continue;
		nativetable = NativeTable()
		nativetable.name_ = sheetname.replace(startswith, "")
		for row_index in range(1, sheet.max_row + 1):
			cols = list()
			for col_index in range(1, sheet.max_column + 1):
				cell = sheet.cell(row = row_index, column = col_index)
				if cell.value:
					cols.append(str(cell.value))
				else:
					cols.append("")
			nativetable.row_list_.append(cols)
		nativetable_list.append(nativetable)
	return nativetable_list

def CreateTable(native_table : NativeTable):
	table = Table()
	table.name_ = native_table.name_
	table.data_count_ = 0
	table.field_dict_ = dict()

	# 맨 앞 셀의 이름을 찾으면 다 넣고 종료.
	# 문자열이 존재하면 해당 필드가 존재하는 것.
	for row in native_table.row_list_:
		rowtype = row[0].lower()
		if rowtype == "field":
			rowdata = row[1:]
			for col_index in range(len(rowdata)):

				# 값이 존재하는 컬럼만.
				value = rowdata[col_index]
				if value:
					# 전체 키워드 검사.
					if value == csharp_strong_keyword_list_ or not IsCorrectName(value):
						print(f"Error!! FieldName='{value}'")
						sys.exit(-1)

					field = Field()
					field.name_ = value
					table.field_dict_[col_index] = field

			break

	# 필드와 동일한 인덱스에서 내용을 다 채워넣음.
	for row in native_table.row_list_:
		rowtype = row[0].lower()
		if rowtype == "type" or rowtype == "comment" or rowtype == "data":
			row_columns = row[1:]
			for col_index in range(len(row_columns)):

				# 필드제목이 있어 필드로 인정되는 컬럼만.
				if col_index in table.field_dict_.keys():
					value = row_columns[col_index]
					field = table.field_dict_[col_index]
					if rowtype == "type":

						# 배열여부 판단하고 배열표시 삭제 (끝이 배열[]로 끝나면 배열).
						type_length = len(value)
						if type_length > 2 and value[type_length - 2] == "[" and value[type_length - 1] == "]":
							value = value[:-2]
							field.is_array_ = True
						
						# 타입명 제외 키워드 검사.
						if value == csharp_weak_keyword_list_ or not IsCorrectName(value):
							print(f"Error!! FieldType='{value}'")
							sys.exit(-1)

						field.type_ = value

						if field.type_ in ["char", "sbyte", "short", "int", "long", "byte", "ushort", "uint", "ulong"]:
							field.kind_ = DataKind.NUMBER
						elif field.type_ in ["float", "double"]:
							field.kind_ = DataKind.REAL
						elif field.type_ in ["bool"]:
							field.kind_ = DataKind.BOOLEAN
						else:
							field.kind_ = DataKind.TEXT
							field.is_array_ = False # 문자열타입은 배열일 수 없음.
							
					elif rowtype == "comment":
						field.comment_ = value
					elif rowtype == "data":
						field.data_.append(value)

	# 데이터 범위 체크....

	# 빈 데이터 기본값 설정 + 최대데이터 카운트 셋팅.	
	for field in table.field_dict_.values():
		data_count = len(field.data_)
		if table.data_count_ < data_count:
			table.data_count_ = data_count

		for data_index in range(data_count):
			value = field.data_[data_index]

			# 값이 존재하지 않을 때.
			if value == "":
				# 정수형.
				if field.kind_ == DataKind.NUMBER:
					field.data_[data_index] = "0"
				# 실수형.
				elif field.kind_ == DataKind.REAL:
					field.data_[data_index] = "0.0"
				# 논리형.
				elif field.kind_ == DataKind.BOOLEAN:
					field.data_[data_index] = "false"
			else:
				if field.kind_ == DataKind.NUMBER:
					field.data_[data_index] = str(int(float(value)))
				elif field.kind_ == DataKind.REAL:
					field.data_[data_index] = str(float(value))
				elif field.kind_ == DataKind.BOOLEAN:
					field.data_[data_index] = "true" if value.lower() == "true" else "false"

	return table


def CreateJsonFileFromTable(file_name : str, table : Table):
	result = str()
	result += "[\n"

	field_count = len(table.field_dict_)
	for row_index in range(table.data_count_):
		result += "\t{\n"
		field_index = 0
		for field in table.field_dict_.values():
			field_name = field.name_
			field_value = DataValueToJsonValue(field.data_[row_index], field.kind_, field.is_array_)
			
			result += f"\t\t\"{field_name}\": {field_value}"
			if field_index + 1 < field_count:
				result += ",\n"
			else:
				result += "\n"
			field_index += 1
		if row_index + 1 < table.data_count_:
			result += "\t},\n"
		else:
			result += "\t}\n"
	result += "]"
	with open(file_name, "w+", encoding = "utf8") as file:
		file.write(result)


def CreateCsFileFromTable(file_name : str, table : Table, make_enum_fields : list):
	table_name = table.name_
	result = str()

	#result += "using DagraacSystems;\n"
	#result += "using DagraacSystems.Table;\n"
	#result += "using DagraacSystems.TableExtension;\n"
	#result += "using System.Collections.Generic;\n"
	result += "\n"
	result += "\n"

	if make_enum_fields != None:
		for make_enum_field in make_enum_fields:
			result += f"public enum {table_name}_{make_enum_field}\n" # CommonTable_Name
			result += "{\n"
			for field in table.field_dict_.values():
				if field.name_ == make_enum_field and (field.kind_ == DataKind.NUMBER or field.kind_ == DataKind.TEXT):
					for data in field.data_:
						if data:
							result += f"\t{data},\n" # 열거체 작성을 원하는 필드의 모든 값은 서로 유니크여야함.
			result += "}\n"
			result += "\n"
			result += "\n"

	# 클래스.
	result += f"public class {table_name}Data : DagraacSystems.ITableData\n"
	result += "{\n"
	for field in table.field_dict_.values():
		field_name = field.name_
		field_type = field.type_
		field_comment = field.comment_
		if field_comment:
			result += f"\t/// <summary>{field_comment}</summary>\n"
		result += f"\tpublic {field_type} {field_name};\n\n"
	#result += "\n"

	#int ToFields();
	result += "\tpublic System.Collections.Generic.List<System.Tuple<string, System.Type, object>> ToFields()\n"
	result += "\t{\n"
	result += "\t\treturn new System.Collections.Generic.List<System.Tuple<string, System.Type, object>>()\n"
	result += "\t\t{\n"
	field_index = 0
	for field in table.field_dict_.values():
		field_name = field.name_
		result += f"\t\t\tnew System.Tuple<string, System.Type, object>(\"{field_name}\", {field_name}.GetType(), {field_name}),\n"
		field_index += 1
	result += "\t\t};\n"
	result += "\t}\n"
	result += "\n"

	#int GetFieldIndex(string name);
	result += "\tpublic int GetFieldIndex(string name)\n"
	result += "\t{\n"
	result += "\t\tswitch (name)\n"
	result += "\t\t{\n"
	field_index = 0
	for field in table.field_dict_.values():
		field_name = field.name_
		result += f"\t\t\tcase \"{field_name}\": return {field_index};\n"
		field_index += 1
	result += "\t\t}\n"
	result += "\t\treturn -1;\n"
	result += "\t}\n"
	result += "\n"

	#string GetFieldName(int index);
	result += "\tpublic string GetFieldName(int index)\n"
	result += "\t{\n"
	result += "\t\tswitch (index)\n"
	result += "\t\t{\n"
	field_index = 0
	for field in table.field_dict_.values():
		field_name = field.name_
		result += f"\t\t\tcase {field_index}: return \"{field_name}\";\n"
		field_index += 1
	result += "\t\t}\n"
	result += "\t\treturn string.Empty;\n"
	result += "\t}\n"
	result += "\n"

	#System.Type GetFieldType(int index);
	result += "\tpublic System.Type GetFieldType(int index)\n"
	result += "\t{\n"
	result += "\t\tswitch (index)\n"
	result += "\t\t{\n"
	field_index = 0
	for field in table.field_dict_.values():
		field_name = field.name_
		field_type = field.type_
		field_comment = field.comment_
		result += f"\t\t\tcase {field_index}: return {field_name}.GetType();\n"
		field_index += 1
	result += "\t\t}\n"
	result += "\t\treturn null;\n"
	result += "\t}\n"
	result += "\n"

	#object GetFieldValue(int index);
	result += "\tpublic object GetFieldValue(int index)\n"
	result += "\t{\n"
	result += "\t\tswitch (index)\n"
	result += "\t\t{\n"
	field_index = 0
	for field in table.field_dict_.values():
		field_name = field.name_
		field_type = field.type_
		field_comment = field.comment_
		result += f"\t\t\tcase {field_index}: return (object){field_name};\n"
		field_index += 1
	result += "\t\t}\n"
	result += "\t\treturn null;\n"
	result += "\t}\n"
	result += "\n"

	#int GetFieldCount();
	result += "\tpublic int GetFieldCount()\n"
	result += "\t{\n"
	field_count = len(table.field_dict_)
	result += f"\t\treturn {field_count};\n"
	result += "\t}\n"
	#result += "\n"
	result += "}"

	with open(file_name, "w+", encoding = "utf8") as file:
		file.write(result)

def OnMain(args):
	print("TableExporter 0.0.4")
	print("made by dagraac")
	print()

	# 필수 인자.
	# 시트를 json과 cs로 익스포트.
	xlsx_file_name = GetAbsPath(str(args[1]))
	json_path = GetAbsPath(str(args[2]))
	cs_path = GetAbsPath(str(args[3]))

	# 추가 인자.
	make_enum_fields = None
	if len(args) > 4:
		make_enum_fields = str(args[4]).split(';')

	BeginTimer(f"\nTotal Time")
	native_table_list = CreateNativeTableListFromXlsxFile(xlsx_file_name)
	for native_table in native_table_list:
		if native_table:
			print(f"{native_table.name_}")
			table = CreateTable(native_table)

			if table:
				json_file_name = f"{json_path}\\{table.name_}.json"
				BeginTimer(f"\t{json_file_name}")
				CreateJsonFileFromTable(json_file_name, table)
				PrintTimer(EndTimer())

				cs_file_name = f"{cs_path}\\{table.name_}Data.cs"
				BeginTimer(f"\t{cs_file_name}")
				CreateCsFileFromTable(cs_file_name, table, make_enum_fields)
				PrintTimer(EndTimer())
			
	PrintTimer(EndTimer())


if __name__ == "__main__":
	OnMain(sys.argv)

# 파일 xlsx json cs

# TableExporter^
#  "D:\Projects\Example\Design\Tables\ExampleTable.xlsx"^
#  "D:\Projects\Example\Client\Assets\Resources\Tables"^
#  "D:\Projects\Example\Client\Assets\Scripts\Tables\Data"

# TableExporter^
#  ".\ExampleTable.xlsx"^
#  "..\..\Client\Assets\Resources\Tables"^
#  "..\..\Client\Assets\Scripts\Tables\Data"